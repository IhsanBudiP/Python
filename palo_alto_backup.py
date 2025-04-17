import logging
import logging.handlers
import os
import datetime
import openpyxl
import requests
import shutil
import rm_old

def backup_paloalto_config(api_url, api_key, backup_dir, device_name, device_ip, wb, sheet, timeout=60):
    """Backup Palo Alto devices."""

    logger = logging.getLogger(__name__)
    logger.setLevel(logging.INFO)
    handler = logging.handlers.SysLogHandler(address='/dev/log', facility=logging.handlers.SysLogHandler.LOG_LOCAL7)
    formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')
    handler.setFormatter(formatter)
    logger.addHandler(handler)

    params = {
        "type": "export",
        "category": "configuration",
        "key": api_key
    }

    timestamp = datetime.datetime.now().strftime("%d-%m-%Y-%H:%M:%S")
    backup_file = os.path.join(backup_dir, f"{device_name}_{timestamp}.xml")

    try:
        response = requests.get(api_url, params=params, verify=False,timeout=timeout)
        response.raise_for_status()

        with open(backup_file, "wb") as f:
            f.write(response.content)

        logger.info(f"Backup konfigurasi {device_name} berhasil dibuat dan disalin ke {backup_file}.")
        sheet.append([device_ip, device_name, "Success", backup_file, timestamp])
        return "Success", backup_file

    except requests.exceptions.RequestException as e:
        logger.error(f"Request Error with {device_ip}: {e}")
        sheet.append([device_ip, device_name, f"Request Error: {e}", "", timestamp])
        return f"Request Error: {e}", None

    except OSError as e:
        logger.error(f"OS Error with {device_ip}: {e}")
        sheet.append([device_ip, device_name, f"OS Error: {e}", "", timestamp])
        return f"OS Error: {e}", None

    except Exception as e:
        logger.error(f"An unexpected error occurred with {device_ip}: {e}")
        sheet.append([device_ip, device_name, f"Unexpected Error: {e}", "", timestamp])
        return f"Unexpected Error: {e}", None

    finally:
        logger.removeHandler(handler)


def backup_multiple_switches(palo_configs, base_backup_dir, excel_base_dir):
    """Backs up multiple palo alto devices and organizes Excel reports by year and month."""
    now = datetime.datetime.now()
    year_str = now.strftime("%Y")
    month_str = now.strftime("%m_%Y")
    day_str = now.strftime("%d_%m_%Y")

    year_excel_dir = os.path.join(excel_base_dir, year_str)
    os.makedirs(year_excel_dir, exist_ok=True)

    excel_file = os.path.join(year_excel_dir, f"backup_report_{month_str}.xlsx")

    try:
        wb = openpyxl.load_workbook(excel_file)
        if day_str in wb.sheetnames:
            sheet = wb[day_str]
        else:
            sheet = wb.create_sheet(day_str)
            sheet.append(["Device IP", "Device Name", "Status", "Backup File", "Timestamp"])
    except FileNotFoundError:
        wb = openpyxl.Workbook()
        sheet = wb.create_sheet(day_str)
        sheet.append(["Device IP", "Device Name", "Status", "Backup File", "Timestamp"])

    for device_ip, config in palo_configs.items():
        device_name = config["name"]
        #site= config["site"]
        api_url = config["api_url"]
        api_key = config["api_key"]
        backup_dir = os.path.join(base_backup_dir, device_name)
        os.makedirs(backup_dir, exist_ok=True)
        backup_paloalto_config(api_url, api_key, backup_dir, device_name, device_ip, wb, sheet)
        rm_old.remove_old_files(backup_dir,90)

    wb.save(excel_file)

    backup_year_excel_dir = os.path.join(backup_excel_dir, year_str) 
    os.makedirs(backup_year_excel_dir, exist_ok=True)
    backup_file = os.path.join(backup_year_excel_dir, f"backup_report_{month_str}.xlsx")
    shutil.copy2(excel_file, backup_file)
    

if __name__ == "__main__":

    palo_configs = {
        "{ip_address}": {
            "name": "{folder_name}",
            #"site": "{location}",
            "api_url": "{api_url}",
            "api_key": "{api_key}"
        },
        "{ip_address}": {
            "name": "{folder_name}",
            #"site": "Yogyakarta",
            "api_url": "{api_url}",
            "api_key": "{api_key}"
        },
    }
    base_backup_dir = "{base_folder}"
    excel_base_dir = "{excel_folder}"
    backup_excel_dir ="{backup_excel_folder}"

    backup_multiple_switches(palo_configs, base_backup_dir, excel_base_dir)