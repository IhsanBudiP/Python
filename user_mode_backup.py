import paramiko
import os
import logging
import logging.handlers
import datetime
import openpyxl
import time
import shutil
import rm_old

def backup_switch_config(device_ip, device_username, device_password, backup_dir, device_name, wb, sheet, timeout=60,switch_exec_password="C0mpn3t!"):
    """Backs up switch config, logs to syslog and adds info to Excel sheet."""

    logger = logging.getLogger(__name__)
    logger.setLevel(logging.INFO)
    handler = logging.handlers.SysLogHandler(address='/dev/log', facility=logging.handlers.SysLogHandler.LOG_LOCAL7)
    formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')
    handler.setFormatter(formatter)
    logger.addHandler(handler)

    timestamp = datetime.datetime.now().strftime("%d-%m-%Y-%H:%M:%S")
    backup_file = os.path.join(backup_dir, f"{device_name}_{timestamp}.cfg")

    try:
        client = paramiko.SSHClient()
        client.set_missing_host_key_policy(paramiko.AutoAddPolicy())
        client.connect(hostname=device_ip, username=device_username, password=device_password, timeout=timeout)

        shell = client.invoke_shell()
        time.sleep(1) 

        shell.send("enable\n")
        time.sleep(1)
        shell.send(f"{switch_exec_password}\n")
        time.sleep(10) 

        shell.send("terminal length 0\n")
        time.sleep(1)

        shell.send("show running-config\n")
        time.sleep(10) 

        output = ""
        while True:
            if shell.recv_ready():
                data = shell.recv(65535).decode('utf-8', errors='ignore')
                output += data
                time.sleep(0.1)
            else:
                break

        with open(backup_file, "w") as outfile:
            outfile.write(output)

        logger.info(f"Backup konfigurasi {device_name} berhasil dibuat dan disalin ke {backup_file}.")
        sheet.append([device_ip, device_name, "Success", backup_file, timestamp])

    except paramiko.SSHException as e:
        logger.error(f"Error backing up {device_ip}: {e}")
        sheet.append([device_ip, device_name, f"SSH Error: {e}", "", timestamp])

    except OSError as e:
        logger.error(f"OS Error with {device_ip}: {e}")
        sheet.append([device_ip, device_name, f"OS Error: {e}", "", timestamp])

    except Exception as e:
        logger.error(f"An unexpected error occurred with {device_ip}: {e}")
        sheet.append([device_ip, device_name, f"Unexpected Error: {e}", "", timestamp])

    finally:
        try:
            client.close() 
        except Exception as e:
            logger.warning(f"Error closing SSH connection to {device_ip}: {e}")
        logger.removeHandler(handler)

def backup_multiple_switches(device_configs, username, password, base_backup_dir, excel_base_dir): 
    """Backs up multiple cisco devices and organizes Excel reports by year and month."""
    now = datetime.datetime.now()
    year_str = now.strftime("%Y")
    month_str = now.strftime("%m_%Y")
    day_str = now.strftime("%d_%m_%Y")

    year_excel_dir = os.path.join(excel_base_dir, year_str) 
    os.makedirs(year_excel_dir, exist_ok=True) 

    excel_file = os.path.join(year_excel_dir, f"backup_report_{month_str}.xlsx") 

    try:
        wb = openpyxl.load_workbook(excel_file)
        if day_str in wb.sheetnames:
            sheet = wb[day_str]
        else:
            sheet = wb.create_sheet(day_str)
            sheet.append(["Device IP", "Device Name", "Status", "Backup File", "Timestamp"])

    except FileNotFoundError:
        wb = openpyxl.Workbook()
        sheet = wb.create_sheet(day_str)
        sheet.append(["Device IP", "Device Name", "Status", "Backup File", "Timestamp"])

    for device_ip, device_name in device_configs.items():
        backup_dir = os.path.join(base_backup_dir, device_name) 
        os.makedirs(backup_dir, exist_ok=True)
        backup_switch_config(device_ip, username, password, backup_dir, device_name, wb, sheet)
        rm_old.remove_old_files(backup_dir,90)

    wb.save(excel_file)

    backup_year_excel_dir = os.path.join(backup_excel_dir, year_str) 
    os.makedirs(backup_year_excel_dir, exist_ok=True)
    backup_file = os.path.join(backup_year_excel_dir, f"backup_report_{month_str}.xlsx")
    shutil.copy2(excel_file, backup_file)

if __name__ == "__main__":

    device_configs = {
        "{ip_address}" : "{folder_name}",
        "{ip_address}" : "{folder_name}",
    }
    username = "{device_username}"
    password = "{device_password}"
    base_backup_dir = "{base_folder}"
    excel_base_dir = "{excel_folder}"
    backup_excel_dir ="{backup_excel_folder}"

    backup_multiple_switches(device_configs, username, password, base_backup_dir, excel_base_dir)